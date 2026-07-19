"""
ResellOS - Agent 10: Buy-Side Stock & Discount Watch
=====================================================
Standalone daily-cadence stock/discount watcher across the 7 retailers
where daily checking is acceptable: Macy's, Target, Kohl's, Walmart,
JCPenney, Christianbook, Amazon. LEGO.com seconds-level restock alerts are
explicitly OUT of scope — Josh already pays for a dedicated service for
that, and general-purpose scraping actors aren't built for that cadence
economically (decided 2026-07-19, see CONTEXT.md).

One watch TARGET per LEGO set (not per set+retailer) — which retailers
actually carry a given set isn't known ahead of time, so each run checks
every active target against every retailer and discovers carriers
naturally. A retailer not carrying a set (found=False) is expected and
normal, not an error — don't treat it as a failure.

Alerts fire on:
  - "restock"            : a set that was out of stock at a retailer on the
                            previous check is now in stock there
  - "discount_<pct>pct"  : current price is >= discount_alert_threshold_pct
                            below msrp (default 20%)
Both can fire on the same check.

--------------------------------------------------------------------------
PER-RETAILER CHECK STATUS AS OF 2026-07-19 — READ BEFORE RELYING ON THIS
--------------------------------------------------------------------------
walmart        PROVEN     Apify e-commerce/walmart-product-detail-scraper.
                           Confirmed live 3x tonight (garbage truck, dune
                           ornithopter, lion knights' castle) — clean
                           priceInfo.price / availability / sellerName
                           fields. Field mapping below is real, not a guess.
kohls          PARTIAL    Apify lexis-solutions/kohls-scraper resolved the
                           correct product page from a search URL, but
                           price/availability came back null on tonight's
                           run. Actor's own advertised success rate is
                           64.3% — treat with suspicion until it's produced
                           a real price at least once.
target         UNVERIFIED Apify bovi/target-products returned 0 items on a
                           search-query input tonight. Walmart only worked
                           once given a direct product URL (found via
                           Firecrawl search first) rather than a search
                           string — try that pattern here before assuming
                           the actor is broken.
macys          UNVERIFIED Apify lexis-solutions/macys-scraper returned 0
                           items on a search URL tonight. Same fix to try
                           first: direct product URL, not a search string.
jcpenney       UNVERIFIED Apify stealth_mode/jcpenney-product-search-scraper
                           FAILED outright (exit code 1) tonight. Needs a
                           different actor or a Firecrawl-only approach.
amazon         DIFFICULT  Apify junglee/Amazon-crawler failed twice tonight
                           — once on a search URL, once on a real, verified
                           direct product URL ("no_results_found" both
                           times). A Firecrawl scrape of the same URL loaded
                           the real page (title matched exactly) but didn't
                           return usable JSON — looks like bot-detection
                           friction, not a bad URL. May need a paid/stealth
                           proxy tier. Confirm the added cost is worth it
                           before investing further — Amazon is one
                           retailer among seven, not the priority.
christianbook  UNVERIFIED No Apify actor exists for this retailer at all —
                           it's a small enough niche that nobody's built
                           one. Firecrawl search found zero results for one
                           test set (10327) tonight, which may just mean
                           Christianbook doesn't carry that particular set
                           — not proof the mechanism doesn't work. Needs a
                           real test against a set Christianbook is known
                           to actually stock.

Bottom line: only Walmart is proven end-to-end tonight. Everything else
needs a first real, verified run before you trust its output. This is
intentional — shipping a guessed field mapping into a tool whose whole
job is "tell me when to act" is worse than shipping nothing.
--------------------------------------------------------------------------

CREDENTIALS THIS SCRIPT NEEDS (neither is set as of 2026-07-19):
  APIFY_API_TOKEN     in .env  — https://console.apify.com/account/integrations
  FIRECRAWL_API_KEY   in .env  — https://www.firecrawl.dev/signin
This script calls the Apify and Firecrawl REST APIs directly, which is a
different auth path than Cowork's already-authenticated MCP connectors
(what tonight's live verification actually used). It will exit with a
clear error if either key is missing — see require_credentials().

Usage: python agents/agent_10_stock_watch.py
"""

import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests

sys.path.insert(0, str(Path(__file__).parent.parent))
from db_client import get_client, PHASE_1_USER_ID

APIFY_API_TOKEN = os.getenv("APIFY_API_TOKEN")
FIRECRAWL_API_KEY = os.getenv("FIRECRAWL_API_KEY")

RETAILERS = ("walmart", "target", "kohls", "macys", "jcpenney", "christianbook", "amazon")

# Apify actor per retailer — see status notes above. No entry for
# christianbook: no actor exists, Firecrawl is the only option there.
APIFY_ACTORS = {
    "walmart":   "e-commerce/walmart-product-detail-scraper",
    "target":    "bovi/target-products",
    "kohls":     "lexis-solutions/kohls-scraper",
    "macys":     "lexis-solutions/macys-scraper",
    "jcpenney":  "stealth_mode/jcpenney-product-search-scraper",
    "amazon":    "junglee/Amazon-crawler",
}

DEFAULT_DISCOUNT_THRESHOLD_PCT = 20


def require_credentials():
    missing = []
    if not APIFY_API_TOKEN:
        missing.append("APIFY_API_TOKEN")
    if not FIRECRAWL_API_KEY:
        missing.append("FIRECRAWL_API_KEY")
    if missing:
        print("ERROR: Missing required .env values: " + ", ".join(missing))
        print("  APIFY_API_TOKEN   — https://console.apify.com/account/integrations")
        print("  FIRECRAWL_API_KEY — https://www.firecrawl.dev/signin")
        sys.exit(1)


# --------------------------------------------------------------------------- #
# Pure logic — no I/O, unit-testable
# --------------------------------------------------------------------------- #

def determine_alert(previous_check, current_check, msrp=None, threshold_pct=DEFAULT_DISCOUNT_THRESHOLD_PCT):
    """
    previous_check / current_check: dicts with at least 'found' and
    'in_stock' keys; current_check also needs 'price' for the discount
    check. Returns a list of reason strings — empty list means no alert.
    A restock and a discount can both fire on the same check.
    """
    reasons = []
    if not current_check.get("found"):
        return reasons  # nothing to alert on if this retailer doesn't carry it

    if previous_check is not None and previous_check.get("found"):
        was_in_stock = previous_check.get("in_stock")
        is_in_stock = current_check.get("in_stock")
        if was_in_stock is False and is_in_stock is True:
            reasons.append("restock")

    price = current_check.get("price")
    if msrp and price is not None and float(msrp) > 0:
        discount_pct = round((float(msrp) - float(price)) / float(msrp) * 100, 1)
        if discount_pct >= threshold_pct:
            reasons.append(f"discount_{discount_pct}pct")

    return reasons


def parse_tier_list_row(set_name_full):
    """
    Parses a Brick Domain tier-list "Set Name" cell (format: "SETNUM Name",
    e.g. "10327 Dune Atreides Royal Ornithopter") into (set_number, name).
    Returns (None, cleaned_string) if no leading set number is found,
    rather than raising — a malformed row shouldn't crash a bulk import.
    """
    text = str(set_name_full).strip()
    m = re.match(r"^(\d{4,6})\s+(.*)$", text)
    if m:
        return m.group(1), m.group(2)
    return None, text


def _first_present(d, keys, default=None):
    """Try several possible dict keys in order, return the first non-None hit."""
    for k in keys:
        v = d.get(k)
        if v is not None:
            return v
    return default


def _coerce_price(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _coerce_in_stock(raw):
    """Normalizes wildly different availability representations across
    retailers into a plain bool, or None if genuinely unknown."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    s = str(raw).strip().upper()
    if s in ("IN_STOCK", "INSTOCK", "IN STOCK", "AVAILABLE", "YES", "TRUE"):
        return True
    if s in ("OUT_OF_STOCK", "OUTOFSTOCK", "OUT OF STOCK", "SOLD OUT", "UNAVAILABLE", "NO", "FALSE"):
        return False
    return None


# --------------------------------------------------------------------------- #
# Apify / Firecrawl REST calls
# --------------------------------------------------------------------------- #

def run_apify_actor(actor_id, input_payload, timeout=120):
    require_credentials()
    url = f"https://api.apify.com/v2/acts/{actor_id.replace('/', '~')}/run-sync-get-dataset-items"
    resp = requests.post(url, params={"token": APIFY_API_TOKEN}, json=input_payload, timeout=timeout)
    resp.raise_for_status()
    return resp.json()


def firecrawl_scrape_json(url, prompt, schema, timeout=60):
    require_credentials()
    resp = requests.post(
        "https://api.firecrawl.dev/v1/scrape",
        headers={"Authorization": f"Bearer {FIRECRAWL_API_KEY}"},
        json={"url": url, "formats": ["json"], "jsonOptions": {"prompt": prompt, "schema": schema}},
        timeout=timeout,
    )
    resp.raise_for_status()
    return resp.json()


def firecrawl_search(query, include_domains=None, limit=3, timeout=30):
    require_credentials()
    payload = {"query": query, "limit": limit}
    if include_domains:
        payload["includeDomains"] = include_domains
    resp = requests.post(
        "https://api.firecrawl.dev/v1/search",
        headers={"Authorization": f"Bearer {FIRECRAWL_API_KEY}"},
        json=payload,
        timeout=timeout,
    )
    resp.raise_for_status()
    return resp.json()


# --------------------------------------------------------------------------- #
# Per-retailer check functions
# --------------------------------------------------------------------------- #
# Each returns a dict: {"found": bool, "in_stock": bool|None, "price": float|None,
# "availability_raw": str|None, "product_url": str|None}. Field-name guesses
# for anything not marked PROVEN above are best-effort — see the module
# docstring before trusting output from an UNVERIFIED retailer.

def check_walmart(target):
    query = f"LEGO {target['set_number']} {target.get('set_name') or ''}".strip()
    search = firecrawl_search(query, include_domains=["walmart.com"], limit=1)
    hits = search.get("data", {}).get("web", [])
    product_urls = [h["url"] for h in hits if "/ip/" in h.get("url", "")]
    if not product_urls:
        return {"found": False, "in_stock": None, "price": None, "availability_raw": None, "product_url": None}

    url = product_urls[0]
    items = run_apify_actor(APIFY_ACTORS["walmart"], {"startUrls": [{"url": url}], "maxProductsPerStartUrl": 1})
    if not items:
        return {"found": False, "in_stock": None, "price": None, "availability_raw": None, "product_url": url}

    item = items[0]
    price_info = item.get("priceInfo", {}) or {}
    availability_raw = item.get("availability")
    return {
        "found": True,
        "in_stock": _coerce_in_stock(availability_raw),
        "price": _coerce_price(price_info.get("price")),
        "availability_raw": availability_raw,
        "product_url": url,
    }


# target / kohls / macys / jcpenney / amazon checkers follow the same shape
# as check_walmart but are UNVERIFIED — see module docstring. Not wired into
# run_daily_check() below yet. Each needs one confirmed live run (direct
# product URL, not a search query — that's the fix Walmart needed too)
# before it should be trusted. Left as a clear next step rather than shipped
# with a guessed field mapping.

RETAILER_CHECKERS = {
    "walmart": check_walmart,
}


# --------------------------------------------------------------------------- #
# Database access
# --------------------------------------------------------------------------- #

def fetch_active_targets(client):
    result = (
        client.table("stock_watch_targets")
        .select("*")
        .eq("user_id", PHASE_1_USER_ID)
        .eq("active", True)
        .execute()
    )
    return result.data or []


def get_last_check(client, target_id, retailer):
    result = (
        client.table("stock_watch_checks")
        .select("*")
        .eq("target_id", target_id)
        .eq("retailer", retailer)
        .order("checked_at", desc=True)
        .limit(1)
        .execute()
    )
    return result.data[0] if result.data else None


def record_check(client, target, retailer, check_result, alert_reasons):
    row = {
        "user_id": PHASE_1_USER_ID,
        "target_id": target["target_id"],
        "retailer": retailer,
        "found": check_result["found"],
        "in_stock": check_result["in_stock"],
        "price": check_result["price"],
        "availability_raw": check_result["availability_raw"],
        "product_url": check_result["product_url"],
        "alert_triggered": bool(alert_reasons),
        "alert_reason": ",".join(alert_reasons) if alert_reasons else None,
    }
    client.table("stock_watch_checks").insert(row).execute()


def run_daily_check(client, retailers=None):
    """Runs every active target against every retailer with a wired-up
    checker (see RETAILER_CHECKERS — only walmart tonight), records each
    check, and prints any alerts."""
    targets = fetch_active_targets(client)
    retailers = retailers or list(RETAILER_CHECKERS.keys())
    alerts = []

    print(f"\nChecking {len(targets)} target(s) across {len(retailers)} retailer(s)...")
    for target in targets:
        for retailer in retailers:
            checker = RETAILER_CHECKERS.get(retailer)
            if checker is None:
                continue
            try:
                result = checker(target)
            except Exception as e:
                print(f"  ERROR checking {target['set_number']} at {retailer}: {e}")
                continue

            previous = get_last_check(client, target["target_id"], retailer)
            reasons = determine_alert(
                previous, result,
                msrp=target.get("msrp"),
                threshold_pct=target.get("discount_alert_threshold_pct") or DEFAULT_DISCOUNT_THRESHOLD_PCT,
            )
            record_check(client, target, retailer, result, reasons)

            if reasons:
                alerts.append((target, retailer, reasons, result))
                print(f"  ALERT: {target['set_number']} {target.get('set_name')} @ {retailer} — {', '.join(reasons)}")
            elif result["found"]:
                print(f"  ok: {target['set_number']} @ {retailer} — in_stock={result['in_stock']} price={result['price']}")

    print(f"\n{len(alerts)} alert(s) this run.")
    return alerts


# --------------------------------------------------------------------------- #
# Target management + tier-list import
# --------------------------------------------------------------------------- #

def import_tier_list_xlsx(client, path, tiers=("High",)):
    """
    Bulk-imports a Brick Domain tier list export (columns: Theme, Set Name,
    Price, Retiring, [Current] Tier, ...) into stock_watch_targets, limited
    to the given tier(s). Skips sets already present (unique on set_number
    per user). Defaults to High tier only — the July 2026 seed used the
    same default; widen to ("High", "High-mid") deliberately, not by
    accident, since more targets means more daily check volume/cost.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Ordered by Tier"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    tier_col = header.index("Current Tier") if "Current Tier" in header else header.index("Tier")

    added, skipped = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        theme, set_name_full, price, retiring = row[0], row[1], row[2], row[3]
        tier = row[tier_col]
        if tier not in tiers:
            continue
        set_number, set_name = parse_tier_list_row(set_name_full)
        if not set_number:
            print(f"  SKIP: could not parse set number from '{set_name_full}'")
            skipped += 1
            continue
        msrp = _coerce_price(price)
        existing = (
            client.table("stock_watch_targets")
            .select("target_id")
            .eq("user_id", PHASE_1_USER_ID)
            .eq("set_number", set_number)
            .execute()
        )
        if existing.data:
            skipped += 1
            continue
        client.table("stock_watch_targets").insert({
            "user_id": PHASE_1_USER_ID,
            "set_number": set_number,
            "set_name": set_name,
            "theme": theme,
            "msrp": msrp,
            "tier": tier,
            "retiring_month": str(retiring).strip() if retiring else None,
        }).execute()
        added += 1

    print(f"Imported {added} target(s), skipped {skipped} (already present or unparseable).")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def list_targets(client):
    targets = fetch_active_targets(client)
    if not targets:
        print("  No active targets.")
        return
    for t in targets:
        print(f"  {t['set_number']:>6}  {t.get('set_name', ''):<40}  tier={t.get('tier')}  msrp=${t.get('msrp')}")


def main():
    require_credentials()
    client = get_client()
    while True:
        print("\n" + "=" * 60)
        print("  RESELLOS -- AGENT 10: BUY-SIDE STOCK & DISCOUNT WATCH")
        print("=" * 60)
        print("  1. List active targets")
        print("  2. Run daily check (wired retailers only — see docstring)")
        print("  3. Import tier list xlsx")
        print("  4. Exit")
        choice = input("Choice [1]: ").strip() or "1"

        if choice == "1":
            list_targets(client)
        elif choice == "2":
            run_daily_check(client)
        elif choice == "3":
            path = input("  Path to tier list .xlsx: ").strip()
            import_tier_list_xlsx(client, path)
        elif choice == "4":
            break
        else:
            print("  Please choose 1-4.")


if __name__ == "__main__":
    main()
